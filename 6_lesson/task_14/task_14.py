from openpyxl import load_workbook
wb=load_workbook('salaries.xlsx')
sheet=wb.active
professions=[]
cities=[]
lst_zp=[]

for i in range(2,sheet.max_column+1):
    professions.append(sheet.cell(row=1,column=i).value)
for i in range(2,sheet.max_row+1):
    cities.append(sheet.cell(row=i,column=1).value)

zp_cities=dict()
zp_professions=dict()
#1 subtask
for i in range(2,sheet.max_row+1):
    for j in range(2,sheet.max_column+1):
        lst_zp.append(sheet.cell(row=i,column=j).value)
    zp_cities[cities[i-2]]=sorted(lst_zp)[len(sorted(lst_zp))//2]
    lst_zp=[]
#2 subtask
for i in range(2,sheet.max_column+1):
    su=0
    for row in sheet.iter_rows(min_row=2, min_col=i,max_col=i, max_row=sheet.max_row, values_only=True):
        su+=row[0]
    zp_professions[professions[i-2]]=su/sheet.max_row-1
print(sorted(zp_cities.items(),key=lambda x: -x[1])[0][0],sorted(zp_professions.items(),key=lambda item: -item[1])[0][0],sep=' ')
