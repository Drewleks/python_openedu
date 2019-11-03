def replace_none(s):
    s=str(s)
    s.replace(',','.')
    if s == 'None':
        s=0
    s=float(s)
    return s

from math import floor

from openpyxl import load_workbook
wb=load_workbook('trekking3.xlsx')
products=dict()
sheet=wb['Справочник']
for i in range(2,sheet.max_row+1):
    products[sheet.cell(row=i,column=1).value]=[replace_none(sheet.cell(row=i,column=2).value),
    replace_none(sheet.cell(row=i,column=3).value),
    replace_none(sheet.cell(row=i,column=4).value),
    replace_none(sheet.cell(row=i,column=5).value)]

sheet=wb['Раскладка']
raskladka=[]
for i in range(2,sheet.max_row+1):
    raskladka.append([sheet.cell(row=i,column=1).value,sheet.cell(row=i,column=2).value,sheet.cell(row=i,column=3).value/100])
#Нам надо найти номер минимального и максиммального дня, чтобы потом сделать range
min_day=sorted(raskladka)[0][0]
max_day=sorted(raskladka,reverse=True)[0][0]
print(min_day,max_day)
for day in range(min_day,max_day+1):
    food_list = []
    for raskl in raskladka:
        if raskl[0]==day:
            food_list.append([raskl[1], raskl[2] * products[raskl[1]][0], raskl[2] * products[raskl[1]][1],
                              raskl[2] * products[raskl[1]][2], raskl[2] * products[raskl[1]][3]])
    kcal = 0#килокалории
    b = 0#белки
    z = 0#жиры
    u = 0#углеводы
    for i in range(0, len(food_list)):
        kcal += food_list[i][1]
        b += food_list[i][2]
        z += food_list[i][3]
        u += food_list[i][4]
    print(*map(floor, (kcal, b, z, u)), sep=' ')