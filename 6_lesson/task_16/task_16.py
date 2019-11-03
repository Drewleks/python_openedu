def replace_none(s):
    s=str(s)
    s.replace(',','.')
    if s == 'None':
        s=0
    s=float(s)

    return s
from math import floor

from openpyxl import load_workbook
wb=load_workbook('trekking2.xlsx')
sheet=wb['Раскладка']
raskladka1=[]
for i in range(2,sheet.max_row+1):
    raskladka1.append([sheet.cell(row=i,column=1).value,sheet.cell(row=i,column=2).value/100])
#print(products)
raskladka=dict()
sheet=wb['Справочник']
for i in range(2,sheet.max_row+1):
    raskladka[sheet.cell(row=i,column=1).value]=[replace_none(sheet.cell(row=i,column=2).value),
    replace_none(sheet.cell(row=i,column=3).value),
    replace_none(sheet.cell(row=i,column=4).value),
    replace_none(sheet.cell(row=i,column=5).value)]
#print(raskladka)
food_list=[]
for item in raskladka1:
    food_list.append([item[0],item[1]*raskladka[item[0]][0],item[1]*raskladka[item[0]][1],
                      item[1]*raskladka[item[0]][2],item[1]*raskladka[item[0]][3]])
for i in food_list:
    print(i)
kcal=0
b=0
z=0
u=0
for i in range(0,len(food_list)):
    kcal+=food_list[i][1]
    b+=food_list[i][2]
    z+=food_list[i][3]
    u+=food_list[i][4]
print(*map(floor,(kcal,b,z,u)),sep=' ')