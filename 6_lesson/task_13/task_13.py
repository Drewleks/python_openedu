import openpyxl
import os

lst=[]
fout=open('out.txt','w',encoding='utf8')
for file in os.listdir("./rogaikopyta"):
    if file.endswith(".xlsx"):
        #print(file)
        wb = openpyxl.load_workbook(os.path.join("./rogaikopyta", file))
        sh = wb.active
        name = sh.cell(row=2, column=2).value
        zp = sh.cell(row=2, column=4).value
        lst.append([name,zp])
for i in sorted(lst):
    print(f'{i[0]} {i[1]}')
    fout.writelines(f'{i[0]} {i[1]}\n')
fout.close()