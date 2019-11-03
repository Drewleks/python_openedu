from openpyxl import load_workbook
wb=load_workbook('trekking1.xlsx')
sheet=wb['Справочник']
d={}
for i in range (2,sheet.max_row+1):
    d[sheet.cell(row=i,column=1).value]=sheet.cell(row=i,column=2).value

fout=open('out.txt','w',encoding='utf8')
for k,v in sorted(d.items(),key=lambda x: (-x[1],x[0])):
    print(k,file=fout)
fout.close()